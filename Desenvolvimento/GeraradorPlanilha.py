import openpyxl
import os
import time

def gerarPlanilha(PRODUTO_INFOS_EXCELL, COLUNAS, PORCENTAGENS):
    try:
        if not os.path.exists("ExcelGerados"):
            os.makedirs("ExcelGerados")
            
        print("\n")
        for contador in range(1, 4):
            print("Gerando Excel" + "." * contador)
            time.sleep(0.35)

        path_arquivos = os.path.join("ExcelGerados", f"{PRODUTO_INFOS_EXCELL['Nome']}.xlsx")
        book = openpyxl.Workbook()

        produto_pagina = book.active  # Obtém a planilha ativa
        linha = 1

        for chave, valor in PRODUTO_INFOS_EXCELL.items():
            produto_pagina.cell(row=linha, column=1, value=chave)  # Adiciona a chave na primeira coluna
            
            if chave != "Nome":
                produto_pagina.cell(row=linha, column=2, value=f"R$ {valor}")  # Adiciona o valor na segunda coluna
            else:
                produto_pagina.cell(row=linha, column=2, value=valor)
            
            produto_pagina.cell(row=linha, column=3,)            
            linha += 1

        j=2
        for i in range(0,len(PORCENTAGENS)):
            #2-9
            produto_pagina.cell(row=1, column=3, value="%")
            produto_pagina.cell(row=j, column=3, value=PORCENTAGENS[i])
            j+=1
            if j > 9:
                break

        book.save(path_arquivos)
    except RuntimeError:
        print(f"\nAlgo deu errado com a criação do arquivo Excel do produto: {PRODUTO_INFOS_EXCELL['Nome']}")
    except PermissionError:
        print("\nErro de permissão.\nVerifique se o arquivo Excel que está sendo gerado não se encontra aberto.")
